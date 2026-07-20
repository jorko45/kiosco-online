"""
K24 - Descargador de logos desde Excel
=======================================
1. Completá logos_pendientes.xlsx con las URLs
2. Ejecutá: python descargar_logos.py
3. Subí index.html + carpeta img/logos/ a GitHub
"""

import os, re, time, urllib.request
from urllib.parse import urlparse

try:
    from openpyxl import load_workbook
except ImportError:
    os.system('pip install openpyxl -q')
    from openpyxl import load_workbook

BASE   = os.path.dirname(__file__)
EXCEL  = os.path.join(BASE, 'logos_pendientes.xlsx')
LOGOS  = os.path.join(BASE, 'img', 'logos')
INDEX  = os.path.join(BASE, 'index.html')

os.makedirs(LOGOS, exist_ok=True)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
    'Referer': 'https://kiosco-online.vercel.app/',
}

wb = load_workbook(EXCEL)
ws = wb.active

with open(INDEX, encoding='utf-8') as f:
    html = f.read()

ok = fail = skip = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    brand_id, brand_name, url = (row + (None,None,None))[:3]
    if not brand_id or not url or str(url).strip() == '':
        continue

    url = str(url).strip()
    ext = os.path.splitext(urlparse(url).path)[1].lower()
    if ext not in ['.jpg','.jpeg','.png','.webp','.svg','.gif']:
        ext = '.png'

    fname    = f"{brand_id}-logo{ext}"
    local    = os.path.join(LOGOS, fname)
    rel_path = f'/img/logos/{fname}'

    if os.path.exists(local) and os.path.getsize(local) > 500:
        print(f"  ⏭  {brand_name} (ya existe)")
        skip += 1
    else:
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = resp.read()
            if len(data) < 300:
                raise ValueError(f"Archivo muy pequeño ({len(data)} bytes)")
            with open(local, 'wb') as f:
                f.write(data)
            print(f"  ✅ {brand_name}")
            ok += 1
        except Exception as e:
            print(f"  ❌ {brand_name}: {e}")
            fail += 1
            continue
        time.sleep(0.3)

    # Actualizar HTML
    import re as _re
    pattern = rf"({{ id:'{_re.escape(brand_id)}'[^}}]*?)logo:(undefined|'[^']*'),"
    new_html = _re.sub(pattern, rf"\1logo:'{rel_path}',", html, flags=_re.DOTALL)
    if new_html != html:
        html = new_html
        print(f"     → HTML actualizado para '{brand_id}'")

with open(INDEX, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\n✅ {ok} descargados  ⏭ {skip} existentes  ❌ {fail} fallidos")
print("📁 Subí img/logos/ e index.html a GitHub")
